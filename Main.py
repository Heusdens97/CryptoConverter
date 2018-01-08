from binance.client import Client
from bitstamp.client import Public
from openpyxl import load_workbook, Workbook
from openpyxl.chart import Reference, LineChart
from openpyxl.chart.axis import DateAxis
from openpyxl.styles import Border, Side, Font
import time
import datetime

init_money = FILL_IN
binance_key = FILL_IN
binance_secret = FILL_IN

binance_client = Client(binance_key, binance_secret)
bitstamp_public_client = Public()

binance_fee = 0.050 / 100
bitstamp_fee = 0.25 / 100
ETH_withdrawal_fee = 0.01000000


def main():
    try:
        wb = load_workbook("Crypto.xlsx")
    except FileNotFoundError:
        wb = Workbook()
        wb.remove_sheet(wb["Sheet"])
        wb.create_sheet("Data")
        wb.create_sheet("Charts")
        wb.create_sheet("Totaal")

    sheet_data = wb["Data"]
    sheet_chart = wb["Charts"]
    wb.remove_sheet(wb["Totaal"])
    sheet_totaal = wb.create_chartsheet("Totaal")

    medium = Side(border_style="medium", color="000000")
    border_right = Border(right=medium)
    border_left = Border(left=medium)
    border_bottom = Border(bottom=medium)
    font = Font(name='Calibri', size = 12)

    sheet_data.cell(column=1, row=1, value="Tijd").border = border_right + border_bottom
    sheet_data.cell(column=1, row=1).font = font
    row_height = 2
    while sheet_data["A" + str(row_height)].value is not None:
        row_height += 1

    ts = time.time()
    sheet_data["A" + str(row_height)] = datetime.datetime.fromtimestamp(ts).strftime('%d-%m-%Y %H:%M:%S')
    sheet_data["A" + str(row_height)].border = border_right
    sheet_data["A" + str(row_height)].font = font
    sheet_data.column_dimensions["A"].width = 18

    column_height = 2
    account = binance_client.get_account()
    tickers = binance_client.get_all_tickers()
    total_eth = 0
    rate_etheur = float(bitstamp_public_client.ticker("eth", "eur")['last'])
    print("asset | amount | EUR")

    timestamp = Reference(sheet_data, min_col=1, min_row=2, max_col=1, max_row=row_height)
    row = 1
    column = "A"
    for b in account['balances']:
        amount = float(b['free'])
        locked = float(b['locked'])
        asset = b['asset']
        if locked > 0:
            print("Locked: " + asset + " " + str(locked))
        if amount > 0:
            sheet_data.cell(column=column_height, row=1, value=asset).border = border_bottom
            sheet_data.cell(column=column_height, row=1, value=asset).font = font
            if asset != "ETH":
                for t in tickers:
                    ticker = asset + "ETH"
                    if ticker == t['symbol']:
                        price = float(t['price'])
                        total_eth += (amount * price*(1-binance_fee))
                        eur = amount * price * rate_etheur
                        break
            else:
                total_eth += amount
                eur = amount * rate_etheur
            eur *= (1-bitstamp_fee)
            print(asset + " | " + str(amount) + " | " + str(eur))
            sheet_data.cell(column=column_height, row=row_height, value=eur).font = font

            values = Reference(sheet_data, min_col=column_height, min_row=2, max_col=column_height, max_row=row_height)
            chart = create_chart(asset, timestamp, values)
            sheet_chart.add_chart(chart, column + str(row))
            row += 14
            if row > 35:
                column = chr(ord(column) + 8)
                row = 1
            column_height += 1

    total_eth -= ETH_withdrawal_fee
    virtual_amount = total_eth*rate_etheur*(1-bitstamp_fee)
    sheet_data.cell(column=column_height, row=1, value="Totaal").border = border_left + border_bottom
    sheet_data.cell(column=column_height, row=1).font = font
    sheet_data.cell(column=column_height, row=row_height, value=virtual_amount).border = border_left
    sheet_data.cell(column=column_height, row=row_height).font = font

    values = Reference(sheet_data, min_col=column_height, min_row=2, max_col=column_height, max_row=row_height)
    chart = create_chart("Totaal", timestamp, values)
    sheet_totaal.add_chart(chart)

    print("Total amount in euro: " + str(virtual_amount))
    print("Profit: " + str(virtual_amount - init_money) + " euro")
    print("ROI: " + str((virtual_amount-init_money)*100/init_money))
    wb.template = False
    wb.save('Crypto.xlsx')


def create_chart(name, x_axis, values):
    chart = LineChart()
    chart.title = name
    chart.x_axis.title = "Tijd"
    chart.x_axis.number_format = '%d-%m-%Y %H:%M:%S'
    chart.x_axis = DateAxis(crossAx=100)
    chart.x_axis.majorTimeUnit = "days"
    chart.y_axis.title = "EUR"
    chart.y_axis.crossAx = 500
    chart.legend = None
    chart.add_data(values)
    chart.set_categories(x_axis)
    return chart


if __name__ == '__main__':
    main()
